from openpyxl import load_workbook
class Book:
    def __init__(self, file_name):
        self.workbook = load_workbook(file_name)

    def sheet(self, attribute):
        return self.workbook[attribute]

book = Book('CAMPOS UTILIZADOS.xlsx')
worksheet = book.sheet('CAMPOS')

for row in worksheet.iter_rows():
    if type(row[0] is None):
        #print('broke')
        pass
    consultant = Consultant(row)
for consultant in Consultant.consultants:
    consultant.reward()