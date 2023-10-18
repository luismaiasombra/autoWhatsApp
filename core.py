from openpyxl import load_workbook
import requests
import sqlite3
class Book:
    def __init__(self, file_name):
        self.workbook = load_workbook(file_name)

    def sheet(self, attribute):
        return self.workbook[attribute]
conn = sqlite3.connect('consultants.db')
cursor = conn.cursor()

cursor.execute('''
    CREATE TABLE IF NOT EXISTS consultants (
        code TEXT,
        name TEXT,
        phone TEXT,
        level TEXT,
        status TEXT,
        debit_status TEXT,
        perks TEXT,
        points REAL,
        points_to_copper REAL,
        points_to_silver REAL,
        points_to_gold REAL,
        points_to_diamond REAL
    )
''')
conn.commit()
conn.close()
class Consultant:
    consultants = []

    def __init__(self, row):
        self.code = row[0].value
        self.name = row[1].value
        self.phone = row[2].value
        self.level = row[3].value
        self.status = row[4].value #current situation of the consultant
        self.debit_status = row[5].value
        self.perks =  row[6].value
        self.points = row[7].value
        self.points_to_copper = row[8].value
        self.points_to_silver = row[9].value
        self.points_to_gold = row[10].value
        self.points_to_diamond = row[11].value
        Consultant.consultants.append(self)
        conn = sqlite3.connect('consultants.db')
        cursor = conn.cursor()

        cursor.execute('''
            INSERT INTO consultants 
            (code, name, phone, level, status, debit_status, perks, points, points_to_copper, points_to_silver, points_to_gold, points_to_diamond)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (self.code, self.name, self.phone, self.level, self.status, self.debit_status, self.perks, self.points, self.points_to_copper, self.points_to_silver, self.points_to_gold, self.points_to_diamond))

        conn.commit()
        conn.close()

        
        
        

    @classmethod
    def print_all(cls):
        for consultant in cls.consultants:
            print("Code:", consultant.code)
            print("Name:", consultant.name)
            print("Phone:", consultant.phone)
            print("Level:", consultant.level)
            print("Status:", consultant.status)
            print("Debit Status:", consultant.debit_status)
            print("Perks:", consultant.perks)
            print("Points:", consultant.points)
            print("Points to Copper:", consultant.points_to_copper)
            print("Points to Silver:", consultant.points_to_silver)
            print("Points to Gold:", consultant.points_to_gold)
            print("Points to Diamond:", consultant.points_to_diamond)
            print("\n")


book = Book('CAMPOS UTILIZADOS.xlsx')
worksheet = book.sheet('CAMPOS')

for row in worksheet.iter_rows():
    Consultant(row)

#Consultant.print_all()



url = "https://v5.chatpro.com.br/chatpro-0e358013c5/api/v1/send_message"

payload = {
    "number": "5588992611377",
    "message": "Hello World!"
}
headers = {
    "accept": "application/json",
    "content-type": "application/json",
    "Authorization": "f7841b1677153acd798304d815db5ef2"
}

response = requests.post(url, json=payload, headers=headers)

print(response.text)


